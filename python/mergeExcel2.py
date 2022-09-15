# -*- coding: cp936 -*-
import os
import sys
import openpyxl

if __name__ == '__main__':
    # print("������·����")
    # path = sys.stdin.readline().strip('\n')

    print("��Ŀ����",
          "�·�", "���ۼ������ɱ�", "�������óɱ�", "���ۼƳɱ�", "���õȱ����ɱ�",
          "�·�", "���ۼ������ɱ�", "�������óɱ�", "���ۼƳɱ�", "���õȱ����ɱ�",
          "�·�", "���ۼ������ɱ�", "�������óɱ�", "���ۼƳɱ�", "���õȱ����ɱ�")
    path = 'D:\pws'
    fn = [filename for filename in os.listdir(path)
          if filename.endswith(('.xlsm')) and filename.startswith('JSXM-')]

    # ���� 1
    month_1 = '202207'
    month_2 = '202212'

    # �½�һ��д��excel
    wb_wt = openpyxl.Workbook()
    filename_wt = 'D:\ͳ�Ʊ���2.xlsx'
    ws_wt = wb_wt.active
    ws_wt.cell(1, 1, "��Ŀ����")
    ws_wt.cell(1, 2, "�·�")
    ws_wt.cell(1, 3, "���ۼ������ɱ�")
    ws_wt.cell(1, 4, "�������óɱ�")
    ws_wt.cell(1, 5, "���ۼƳɱ�")
    ws_wt.cell(1, 6, "���õȱ����ɱ�")
    ws_wt.cell(1, 7, "�·�")
    ws_wt.cell(1, 8, "���ۼ������ɱ�")
    ws_wt.cell(1, 9, "�������óɱ�")
    ws_wt.cell(1, 10, "���ۼƳɱ�")
    ws_wt.cell(1, 11, "���õȱ����ɱ�")
    ws_wt.cell(1, 12, "�·�")
    ws_wt.cell(1, 13, "���ۼ������ɱ�")
    ws_wt.cell(1, 14, "�������óɱ�")
    ws_wt.cell(1, 15, "���ۼƳɱ�")
    ws_wt.cell(1, 16, "���õȱ����ɱ�")

    # ���� 2�� filename
    row_wt = 2
    for filename in fn:
        file = path + "\\" + filename
        split_filename_index = filename.index('_')
        project_name = filename[0:split_filename_index]
        print(file)
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb['�¶�Ԥ������ͳ��']
        columnIndex = 0
        for columnNumber in range(3, ws.max_column + 1):
            if ws.cell(row=2, column=columnNumber).value == month_1:
                index_column_1 = columnNumber
            if ws.cell(row=2, column=columnNumber).value == month_2:
                index_column_2 = columnNumber
            columnIndex = columnIndex + 1

        for i in range(3, columnIndex):
            if ws.cell(4, i).value == None:
                index_column_last = i - 1
                break

        month_one = ws.cell(2, index_column_1).value
        human_cost_one = ws.cell(4, index_column_1).value
        other_cost_one = ws.cell(6, index_column_1).value
        accu_cost_one = ws.cell(7, index_column_1).value

        month_two = ws.cell(2, index_column_2).value
        human_cost_two = ws.cell(4, index_column_2).value
        other_cost_two = ws.cell(6, index_column_2).value
        accu_cost_two = ws.cell(7, index_column_2).value

        month_last = ws.cell(2, index_column_last).value
        human_cost_last = ws.cell(4, index_column_last).value
        other_cost_last = ws.cell(6, index_column_last).value
        accu_cost_last = ws.cell(7, index_column_last).value

        # 2 ��ȡ���ñ����ɱ���������Ϊ�ѱ������ã�
        ws2 = wb['���ü������ɱ�']
        for columnNumber in range(5, ws2.max_column + 1):
            tmp_value = ws2.cell(9, columnNumber).value
            if tmp_value == None:
                index_column_last = columnNumber - 1
                break
            if str(tmp_value.strftime("%Y%m")) == month_1:
                index_column_1 = columnNumber
            if str(tmp_value.strftime("%Y%m")) == month_2:
                index_column_2 = columnNumber

        row_list = list(range(10, 19))
        row_list += list(range(23, 27))
        row_list += list(range(32, 36))
        row_list += list(range(41, 44))
        row_list += list(range(49, 53))
        row_list += list(range(58, 61))

        sum_one = 0
        for row in row_list:
            for column in range(5, index_column_1 + 1):
                cell = ws2.cell(row, column)
                if cell.data_type == 'n' and cell.value != None and cell.font.bold == True:
                    sum_one += cell.value

        sum_two = 0
        for row in row_list:
            for column in range(index_column_1 + 1, index_column_2 + 1):
                cell = ws2.cell(row, column)
                if cell.data_type == 'n' and cell.value != None and cell.font.bold == True:
                    sum_two += cell.value

        sum_three = 0
        for row in row_list:
            for column in range(index_column_2 + 1, index_column_last + 1):
                cell = ws2.cell(row, column)
                if cell.data_type == 'n' and cell.value != None and cell.font.bold == True:
                    sum_two += cell.value
        # print(project_name,
        #       month_one, human_cost_one, other_cost_one, accu_cost_one, sum_one,
        #       month_two, human_cost_two, other_cost_two, accu_cost_two, sum_two,
        #       month_last, human_cost_last, other_cost_last, accu_cost_last, sum_three)

        ws_wt.cell(row_wt, 1, project_name)
        ws_wt.cell(row_wt, 2, month_one)
        ws_wt.cell(row_wt, 3, human_cost_one)
        ws_wt.cell(row_wt, 4, other_cost_one)
        ws_wt.cell(row_wt, 5, accu_cost_one)
        ws_wt.cell(row_wt, 6, sum_one)
        ws_wt.cell(row_wt, 7, month_two)
        ws_wt.cell(row_wt, 8, human_cost_two)
        ws_wt.cell(row_wt, 9, other_cost_two)
        ws_wt.cell(row_wt, 10, accu_cost_two)
        ws_wt.cell(row_wt, 11, sum_two)
        ws_wt.cell(row_wt, 12, month_last)
        ws_wt.cell(row_wt, 13, human_cost_last)
        ws_wt.cell(row_wt, 14, other_cost_last)
        ws_wt.cell(row_wt, 15, accu_cost_last)
        ws_wt.cell(row_wt, 16, sum_three)
        row_wt += 1

        index_column_1 = 0
        index_column_2 = 0
        index_column_3 = 0
        project_name = 0
        month_one= 0
        human_cost_one= 0
        other_cost_one= 0
        accu_cost_one= 0
        sum_one= 0
        month_two= 0
        human_cost_two= 0
        other_cost_two= 0
        accu_cost_two= 0
        sum_two= 0
        month_last= 0
        human_cost_last= 0
        other_cost_last= 0
        accu_cost_last= 0
        sum_three= 0
    wb_wt.save(filename_wt)